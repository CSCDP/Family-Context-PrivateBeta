import argparse
import yaml
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment


def parse_openapi(filename):
    with open(filename, 'rt') as FILE:
        spec = yaml.safe_load(FILE)

    schema = spec["components"]["schemas"]

    result_list = []
    for name, definition in schema.items():
        for prop, prop_def in definition.get("properties", {}).items():
            result = dict(
                object=name,
                object_desc=definition.get('description', ''),
                object_name=definition.get('title', ''),
                field=prop
            )
            result.update(prop_def)
            result_list.append(result)

    return result_list


def api_spec_to_excel(spec, dest_filename):
    df = pd.DataFrame(spec)
    df.index = pd.MultiIndex.from_frame(
        df[["object","object_desc","object_name"]],
        names=["Object", "Description", "Name"]
    )
    del df["object"]
    del df["object_desc"]
    del df["object_name"]

    df = df[["field","title","description","type","format","pattern","$ref","items"]]

    sheet_name = "Family Context API Objects"

    writer = pd.ExcelWriter(dest_filename, engine='openpyxl')
    writer.book = Workbook()
    sheet = writer.book.active
    sheet.title = sheet_name

    # Set width
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['E'].width = 25
    sheet.column_dimensions['F'].width = 25
    sheet.column_dimensions['I'].width = 16
    sheet.column_dimensions['J'].width = 30
    sheet.column_dimensions['K'].width = 35


    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    df.to_excel(writer, sheet_name, index=True, merge_cells=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = None
            cell.alignment = Alignment(vertical='top')

        row[1].alignment = Alignment(wrap_text=True)
        row[5].alignment = Alignment(wrap_text=True)


    writer.save()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Convert OpenAPI Schema to excel report')
    parser.add_argument('schema', metavar='<filename>', type=str,
                        help='Filename of the input schema')
    parser.add_argument('output', metavar='<output>', type=str,
                        help='Filename of where to write the output Excel summary')

    args = parser.parse_args()

    api_spec = parse_openapi(args.schema)
    api_spec_to_excel(api_spec, args.output)
